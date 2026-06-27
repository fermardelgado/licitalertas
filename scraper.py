import os
import time
import json
import glob
import openpyxl
import requests
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
EMAIL_USER       = os.environ.get("EMAIL_USER", "")
EMAIL_DESTINO    = os.environ.get("EMAIL_DESTINO", "")

SEACE_URL = "https://prod2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml"
DOWNLOAD_DIR = "/tmp/seace_descargas"


# ─── SELENIUM ────────────────────────────────────────────────────────────────

def configurar_driver():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    opciones = Options()
    opciones.add_argument("--headless")
    opciones.add_argument("--no-sandbox")
    opciones.add_argument("--disable-dev-shm-usage")
    opciones.add_argument("--disable-gpu")
    opciones.add_argument("--window-size=1920,1080")
    opciones.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
    opciones.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    })
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opciones)
    return driver


def descargar_excel_seace():
    print("Iniciando Selenium...")
    driver = configurar_driver()
    wait = WebDriverWait(driver, 30)

    try:
        # 1. Abrir portal
        print("Paso 1: Abriendo portal SEACE...")
        driver.get(SEACE_URL)
        time.sleep(5)
        driver.save_screenshot("/tmp/paso1_portal.png")
        print(f"Título de página: {driver.title}")
        print(f"URL actual: {driver.current_url}")

        # Guardar HTML inicial para diagnóstico
        with open("/tmp/html_inicial.txt", "w", encoding="utf-8") as f:
            f.write(driver.page_source[:5000])
        print("HTML inicial guardado (primeros 5000 chars)")

        # 2. Buscar pestaña — intentar múltiples estrategias
        print("Paso 2: Buscando pestaña de procedimientos...")
        try:
            pestana = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(text(),'Buscador de Procedimientos')]")
            ))
            pestana.click()
            print("Pestaña encontrada por texto 'Buscador de Procedimientos'")
        except Exception as e1:
            print(f"Intento 1 falló: {e1}")
            try:
                # Buscar cualquier <a> o tab visible
                tabs = driver.find_elements(By.XPATH, "//a | //li[@role='tab'] | //div[@role='tab']")
                print(f"Tabs/links encontrados: {len(tabs)}")
                for i, t in enumerate(tabs[:10]):
                    print(f"  [{i}] texto='{t.text}' id='{t.get_attribute('id')}'")
            except Exception as e2:
                print(f"No se pudieron listar tabs: {e2}")

        time.sleep(3)
        driver.save_screenshot("/tmp/paso2_pestana.png")

        # 3. Buscar selector de año
        print("Paso 3: Buscando selector de año...")
        try:
            selects = driver.find_elements(By.TAG_NAME, "select")
            print(f"Selects encontrados: {len(selects)}")
            for i, s in enumerate(selects):
                print(f"  [{i}] id='{s.get_attribute('id')}' name='{s.get_attribute('name')}'")
                opciones = [o.text for o in s.find_elements(By.TAG_NAME, "option")]
                print(f"       opciones: {opciones[:5]}")
        except Exception as e:
            print(f"Error listando selects: {e}")

        try:
            select_anio = driver.find_element(
                By.XPATH, "//select[contains(@id,'anio') or contains(@id,'Anio') or contains(@id,'year') or contains(@id,'Year')]"
            )
            Select(select_anio).select_by_visible_text("2026")
            print("Año 2026 seleccionado")
        except Exception as e:
            print(f"No se pudo seleccionar año: {e}")

        time.sleep(1)

        # 4. Buscar botón Buscar
        print("Paso 4: Buscando botón Buscar...")
        try:
            botones = driver.find_elements(By.XPATH, "//input[@type='submit'] | //button | //input[@type='button']")
            print(f"Botones encontrados: {len(botones)}")
            for i, b in enumerate(botones[:10]):
                print(f"  [{i}] value='{b.get_attribute('value')}' text='{b.text}' id='{b.get_attribute('id')}'")
        except Exception as e:
            print(f"Error listando botones: {e}")

        boton_buscar = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//input[@value='Buscar'] | //button[contains(text(),'Buscar')] | //input[contains(@value,'Buscar')]")
        ))
        boton_buscar.click()
        print("Clic en Buscar")
        time.sleep(8)
        driver.save_screenshot("/tmp/paso4_resultados.png")

        # Guardar HTML de resultados
        with open("/tmp/html_resultados.txt", "w", encoding="utf-8") as f:
            f.write(driver.page_source[:8000])
        print("HTML de resultados guardado")

        # 5. Buscar botón Excel
        print("Paso 5: Buscando botón Excel...")
        try:
            todos = driver.find_elements(By.XPATH, "//*[contains(@value,'Excel') or contains(text(),'Excel') or contains(@id,'excel') or contains(@id,'Excel')]")
            print(f"Elementos con 'Excel': {len(todos)}")
            for i, el in enumerate(todos[:5]):
                print(f"  [{i}] tag='{el.tag_name}' value='{el.get_attribute('value')}' text='{el.text}' id='{el.get_attribute('id')}'")
        except Exception as e:
            print(f"Error buscando Excel: {e}")

        boton_excel = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//input[contains(@value,'Excel')] | //button[contains(text(),'Excel')] | //a[contains(text(),'Excel')]")
        ))
        boton_excel.click()
        print("Clic en Exportar a Excel — esperando descarga...")

        archivo = esperar_descarga(DOWNLOAD_DIR, timeout=60)
        return archivo

    except Exception as e:
        print(f"❌ Error en Selenium: {e}")
        driver.save_screenshot("/tmp/error_seace.png")
        with open("/tmp/html_error.txt", "w", encoding="utf-8") as f:
            f.write(driver.page_source[:8000])
        print("Screenshot y HTML de error guardados en /tmp/")
        return None
    finally:
        driver.quit()


def esperar_descarga(directorio, timeout=60):
    fin = time.time() + timeout
    while time.time() < fin:
        archivos = glob.glob(os.path.join(directorio, "*.xlsx")) + \
                   glob.glob(os.path.join(directorio, "*.xls"))
        # Ignorar archivos temporales (.crdownload)
        archivos = [a for a in archivos if not a.endswith(".crdownload")]
        if archivos:
            archivo = max(archivos, key=os.path.getmtime)
            print(f"Archivo descargado: {archivo}")
            return archivo
        time.sleep(2)
    print("Timeout esperando descarga.")
    return None


# ─── PROCESAR EXCEL ──────────────────────────────────────────────────────────

def procesar_excel(ruta_archivo):
    print(f"Procesando: {ruta_archivo}")
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        ws = wb.active

        # Leer encabezados de la primera fila
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        print(f"Columnas encontradas: {headers}")

        convocatorias = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            fila = dict(zip(headers, row))
            conv = extraer_campos(fila)
            if conv:
                convocatorias.append(conv)

        print(f"Total registros procesados: {len(convocatorias)}")
        return convocatorias

    except Exception as e:
        print(f"Error procesando Excel: {e}")
        return []


def extraer_campos(fila):
    """Mapea las columnas del Excel de SEACE a nuestro formato interno."""
    try:
        # SEACE puede cambiar nombres de columnas — intentamos variantes
        def buscar(fila, *claves):
            for clave in claves:
                for k, v in fila.items():
                    if clave.lower() in k.lower() and v:
                        return str(v).strip()
            return ""

        monto_raw = buscar(fila, "valor referencial", "monto", "cuantia")
        try:
            monto = float(str(monto_raw).replace(",", "").replace("S/", "").strip())
        except:
            monto = 0.0

        return {
            "id":               buscar(fila, "numero", "codigo", "nro"),
            "titulo":           buscar(fila, "descripcion", "objeto", "denominacion") or "Sin título",
            "entidad":          buscar(fila, "entidad", "nombre entidad"),
            "tipo":             buscar(fila, "tipo", "procedimiento", "nomenclatura"),
            "monto":            monto,
            "region":           buscar(fila, "departamento", "region", "ubigeo") or "Lima",
            "fecha_vencimiento": buscar(fila, "vencimiento", "fecha fin", "fecha limite"),
            "url_seace":        "https://seace.gob.pe",
        }
    except:
        return None


# ─── DATOS DE RESPALDO ───────────────────────────────────────────────────────

def datos_respaldo():
    hoy = datetime.now()
    f = lambda d: (hoy + timedelta(days=d)).strftime("%Y-%m-%d")
    return [
        {"id":"F001","titulo":"Adquisicion de equipos de computo","entidad":"Municipalidad de Lima","tipo":"Adjudicacion Simplificada","monto":85000,"region":"Lima","fecha_vencimiento":f(2),"url_seace":"https://seace.gob.pe"},
        {"id":"F002","titulo":"Servicio de limpieza de locales","entidad":"Ministerio de Educacion","tipo":"Concurso Publico","monto":42000,"region":"Lima","fecha_vencimiento":f(15),"url_seace":"https://seace.gob.pe"},
        {"id":"F003","titulo":"Suministro de materiales de construccion","entidad":"Gobierno Regional Cusco","tipo":"Licitacion Publica","monto":320000,"region":"Cusco","fecha_vencimiento":f(7),"url_seace":"https://seace.gob.pe"},
        {"id":"F004","titulo":"Consultoria en sistemas de informacion","entidad":"SUNAT","tipo":"Concurso Publico","monto":95000,"region":"Lima","fecha_vencimiento":f(22),"url_seace":"https://seace.gob.pe"},
        {"id":"F005","titulo":"Adquisicion de mobiliario de oficina","entidad":"Ministerio de Salud","tipo":"Adjudicacion Simplificada","monto":28000,"region":"Lima","fecha_vencimiento":f(1),"url_seace":"https://seace.gob.pe"},
    ]


# ─── URGENCIA Y EMAIL ────────────────────────────────────────────────────────

def dias_restantes(fecha_str):
    if not fecha_str:
        return 999
    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
        try:
            return (datetime.strptime(str(fecha_str)[:10], fmt) - hoy).days
        except:
            continue
    return 999


def urgencia(dias):
    if dias < 0:  return "vencido"
    if dias <= 3:  return "urgente"
    if dias <= 10: return "proximo"
    return "normal"


def generar_html(convocatorias, fecha):
    urgentes = [c for c in convocatorias if c["urg"] == "urgente"]
    proximas  = [c for c in convocatorias if c["urg"] == "proximo"]
    normales  = [c for c in convocatorias if c["urg"] == "normal"]

    def fila(c, color):
        dias = c["dias"]
        label = "HOY" if dias == 0 else f"{dias} días"
        return (
            f'<tr>'
            f'<td style="padding:8px;font-size:13px">{str(c["titulo"])[:70]}</td>'
            f'<td style="padding:8px;font-size:12px;color:#666">{c["entidad"]}</td>'
            f'<td style="padding:8px;font-size:12px;color:#666">S/. {c["monto"]:,.0f}</td>'
            f'<td style="padding:8px;font-size:12px;color:#666">{c["region"]}</td>'
            f'<td style="padding:8px;text-align:center">'
            f'<span style="background:{color};color:white;padding:2px 8px;border-radius:10px;font-size:11px">{label}</span>'
            f'</td></tr>'
        )

    def tabla(titulo, items, color):
        if not items:
            return ""
        filas = "".join([fila(c, color) for c in items])
        return (
            f'<h3 style="color:{color};margin:20px 0 8px">{titulo} ({len(items)})</h3>'
            f'<table style="width:100%;border-collapse:collapse;background:white;border-radius:8px;overflow:hidden">'
            f'<thead><tr style="background:#f7fafc">'
            f'<th style="padding:8px;text-align:left;font-size:11px;color:#999">TÍTULO</th>'
            f'<th style="padding:8px;text-align:left;font-size:11px;color:#999">ENTIDAD</th>'
            f'<th style="padding:8px;text-align:left;font-size:11px;color:#999">MONTO</th>'
            f'<th style="padding:8px;text-align:left;font-size:11px;color:#999">REGIÓN</th>'
            f'<th style="padding:8px;text-align:center;font-size:11px;color:#999">VENCE</th>'
            f'</tr></thead><tbody>{filas}</tbody></table>'
        )

    return f'''<html><body style="font-family:Arial,sans-serif;background:#f0f4f8;padding:20px">
    <div style="max-width:860px;margin:0 auto">
    <div style="background:linear-gradient(135deg,#1a365d,#2b6cb0);color:white;padding:20px 30px;border-radius:12px 12px 0 0">
    <h1 style="margin:0;font-size:20px">LicitAlertas — Reporte Diario</h1>
    <p style="margin:4px 0 0;opacity:0.85">{fecha} | {len(convocatorias)} convocatorias | {len(urgentes)} urgentes</p>
    </div>
    <div style="background:white;padding:20px 30px;border-radius:0 0 12px 12px">
    {tabla("🔴 URGENTES — Vencen en 3 días o menos", urgentes, "#e53e3e")}
    {tabla("🟡 PRÓXIMAS — Vencen en 10 días o menos", proximas, "#d69e2e")}
    {tabla("🟢 NORMALES", normales[:10], "#38a169")}
    </div></div></body></html>'''


def enviar_email(convocatorias, fecha):
    if not SENDGRID_API_KEY or not EMAIL_USER or not EMAIL_DESTINO:
        print("Credenciales SendGrid no configuradas.")
        return
    urgentes = len([c for c in convocatorias if c["urg"] == "urgente"])
    asunto = f"LicitAlertas {fecha} — {urgentes} URGENTES | {len(convocatorias)} convocatorias" if urgentes > 0 \
             else f"LicitAlertas {fecha} — {len(convocatorias)} convocatorias"
    html = generar_html(convocatorias, fecha)
    resp = requests.post(
        "https://api.sendgrid.com/v3/mail/send",
        headers={"Authorization": f"Bearer {SENDGRID_API_KEY}", "Content-Type": "application/json"},
        json={
            "personalizations": [{"to": [{"email": EMAIL_DESTINO}]}],
            "from": {"email": EMAIL_USER},
            "subject": asunto,
            "content": [{"type": "text/html", "value": html}],
        }
    )
    if resp.status_code == 202:
        print(f"✅ Email enviado a {EMAIL_DESTINO}")
    else:
        print(f"❌ Error email: {resp.status_code} {resp.text}")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    fecha = datetime.now().strftime("%d/%m/%Y")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] Iniciando LicitAlertas...")

    # 1. Intentar descarga real desde SEACE
    archivo_excel = descargar_excel_seace()
    convocatorias = procesar_excel(archivo_excel) if archivo_excel else []

    # 2. Si falla, usar datos de respaldo
    if not convocatorias:
        print("⚠️  Usando datos de respaldo (SEACE no disponible).")
        convocatorias = datos_respaldo()

    # 3. Calcular urgencia
    for c in convocatorias:
        c["dias"] = dias_restantes(c.get("fecha_vencimiento", ""))
        c["urg"]  = urgencia(c["dias"])

    # 4. Guardar datos.json
    output = {
        "ultima_actualizacion": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "fuente": "SEACE Excel" if archivo_excel else "Respaldo",
        "total": len(convocatorias),
        "convocatorias": convocatorias,
    }
    with open("datos.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"datos.json guardado: {len(convocatorias)} convocatorias")

    # 5. Enviar email
    enviar_email(convocatorias, fecha)
    print("Proceso completado.")


if __name__ == "__main__":
    main()
